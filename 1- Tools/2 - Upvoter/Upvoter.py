from requests.auth import HTTPProxyAuth
from requests import Session
import praw
import time



def reddit_upvote(url_list, username, password, PROXY_IP_and_PORT, PROXY_USER, PROXY_PASS, client_id, client_secret, user_agent):

    session = Session()
    proxy_phrase = "http://" + PROXY_USER + ":" + PROXY_PASS + "@" + PROXY_IP_and_PORT

    # print(proxy_phrase)

    # session.proxies['http'] = 'http://j3ety:3gf7nsbm@104.153.81.68:5432'

    session.proxies['http'] = str(proxy_phrase)

    ext_ip = session.get('http://checkip.dyndns.org')
    print (ext_ip.text)

    # print(client_id,client_secret,user_agent)

    reddit = praw.Reddit(client_id=client_id,
                         client_secret=client_secret,
                         password=password,
                         requestor_kwargs={'session': session},  # pass Session
                         user_agent=user_agent,
                         username=username)

    ext_ip = session.get('http://checkip.dyndns.org')
    # print (ext_ip.text)

    # print(reddit.user.me())
    # print(reddit._core._requestor._http.proxies)

    # print(reddit.user.me())

    # url iteration
    for url in url_list:
        print(url)

        response = session.get(url)
        print("URL Response: ", response)

        print('Post located.. sleeping')
        time.sleep(50)

        needed_str = ""
        str_list = url.split("/")
        for index, string in enumerate(str_list):
            if string == "comments":
                needed_str = str_list[index + 1]
                # print(needed_str)
                break
            else:
                needed_str = ""

        # submission = reddit.submission(id="n6b2tp")
        submission = reddit.submission(id=needed_str)
        submission.upvote()
        print('Upvoted Successfully.. Sleeping again')
        time.sleep(50)

    # Closes all adapters and as such the session
    session.close()

    time.sleep(25)

    print("switching to another account")

    # submission = reddit.submission(id="n6b2tp")
    # submission.downvote()


#############################################################################################


###################### read "Links.txt" file #############################################

url_list = []  # create empty list for urls

f = open("Links.txt", "r")  # file read method

# contents =f.read()         # read the whole file and store in "contents" variable

line = f.readline()
while line != "":
    url_list.append(line)
    # print(line)
    line = f.readline()

f.close()

###########################################################################################

###################### read "accounts.xlsx" file ##########################################
# import openpyxl module
import openpyxl

user_list = []
pass_list = []
proxy_ip = []
proxy_port = []
proxy_user = []
proxy_pass = []

client_id_list = []
client_secret_list = []
user_agent_list = []

# Give the location of the file
path = "accounts.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row  # get the maximum rows

# print("max rows = ",m_row)

# Loop will print all values
# of first column
for i in range(2, m_row + 1):
    username = sheet_obj.cell(row=i, column=1)
    password = sheet_obj.cell(row=i, column=2)
    proxy_data = sheet_obj.cell(row=i, column=3)

    client_id = sheet_obj.cell(row=i, column=4)
    client_secret = sheet_obj.cell(row=i, column=5)
    user_agent = sheet_obj.cell(row=i, column=6)

    # print(username.value)
    # print(password.value)
    # print(proxy_data.value)
    # print(client_id.value)
    # print(client_secret.value)
    # print(user_agent.value)

    # spliting proxy into ip , port , use and pass
    if proxy_data.value != "":
        proxy_list = proxy_data.value.split(":")
    else:
        string = ":::"
        proxy_list = string.split(":")

    user_list.append(username.value)
    pass_list.append(password.value)

    proxy_ip.append(proxy_list[0])
    proxy_port.append(proxy_list[1])
    proxy_user.append(proxy_list[2])
    proxy_pass.append(proxy_list[3])

    client_id_list.append(client_id.value)
    client_secret_list.append(client_secret.value)
    user_agent_list.append(user_agent.value)

    # print()

###########################################################################################
######################################### Main ############################################
# for loop over accounts
for i, user in enumerate(user_list):
    user_name = user
    pass_word = pass_list[i]
    PROXY_IP = proxy_ip[i]
    PROXY_PORT = proxy_port[i]
    PROXY_USER = proxy_user[i]
    PROXY_PASS = proxy_pass[i]

    CLIENT_ID = client_id_list[i]
    CLIENT_SECRET = client_secret_list[i]
    USER_AGENT = user_agent_list[i]

    PROXY_IP_and_PORT = PROXY_IP + ":" + PROXY_PORT

    # print(user_name , pass_word)
    print("user_name: ", user_name)
    reddit_upvote(url_list, user_name, pass_word, PROXY_IP_and_PORT, PROXY_USER, PROXY_PASS, CLIENT_ID, CLIENT_SECRET,
                  USER_AGENT)

print("Program ran successfully")

